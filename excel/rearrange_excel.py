from openpyxl import load_workbook
from openpyxl import Workbook


def call():
    inpfile = "/home/navendu/Downloads/inp.xlsx"
    opfile = "/home/navendu/Downloads/op.xlsx"

    inpbook = load_workbook(inpfile)
    opbook = load_workbook(opfile)
    sheet_name = 'sheet 1'
    inpsheet = inpbook.get_sheet_by_name(sheet_name)
    opsheet = opbook.get_sheet_by_name(sheet_name)

    row = 3
    for i in inpsheet.columns[1]:
        if '_x000D_' in i.value:
            items = i.value.split("_x000D_")

            name = items[0].strip()
            phone = items[1].strip()
            email = items[2].strip()
            address = ", ".join(items[3:len(items)-1])

            xyz = items[len(items)-1].strip().split()
            city = " ".join(xyz[:len(xyz)-2])
            state = xyz[len(xyz)-2].strip() 
            zp = xyz[len(xyz)-1].strip()

        else:
            items = i.value.split("\n")

            name = items[0].strip()
            phone = items[1].strip()
            email = items[2].strip()
            address = ", ".join(items[3:len(items)-1])

            xyz = items[len(items)-1].strip().split()
            city = " ".join(xyz[:len(xyz)-2])
            state = xyz[len(xyz)-2].strip() 
            zp = xyz[len(xyz)-1].strip()

        opsheet.cell(row=row, column=1).value = name
        opsheet.cell(row=row, column=2).value = address
        opsheet.cell(row=row, column=3).value = city
        opsheet.cell(row=row, column=4).value = state
        opsheet.cell(row=row, column=5).value = zp.replace("'", "")
        opsheet.cell(row=row, column=6).value = phone.replace("'", "")
        opsheet.cell(row=row, column=7).value = email

        if row == 23:
            break
        row += 1

    opbook.save(opfile)





if __name__ == "__main__":
    call()
