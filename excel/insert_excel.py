from time import sleep
import json
import urllib
import urllib2
from openpyxl import load_workbook
from openpyxl import Workbook
import webbrowser

# ADD SERIAL NUMBER (1 or 10501)
# combine ing and ing count!

def insert_excel():

    file_path = '/home/navendu/Downloads/recipes_351_700.xlsx'
    book = load_workbook(file_path)
    sheet_name = 'recipe_bank'
    sheet = book.get_sheet_by_name(sheet_name)

    page_count = 351
    page_max = 701
    while page_count != page_max:

        url ="http://food2fork.com/api/search"
        query_args = {"key": "e1254d39d79328ef647ae25737895fe8",
                      "page": page_count}

        encoded_args = urllib.urlencode(query_args)
        complete_url = url + "?" + encoded_args
        resp = urllib2.urlopen(complete_url)
        data = json.load(resp)

        if len(data['recipes']) != 30:
            print "Num records %s for page %s" % (len(data['recipes']), page_count)

        for i in data['recipes']:
            sheet.append([i.get("publisher", ""),
                          i.get("f2f_url", ""),
                          i.get("title", ""),
                          i.get("source_url", ""),
                          i.get("recipe_id", "").replace("'", ""),
                          i.get("image_url", ""),
                          i.get("social_rank", "").replace("'", ""),
                          i.get("publisher_url", "")])

        page_count += 1
        sleep(1)

    book.save(file_path)

    


    """ NEW EXCEL FILE
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook('demo.xlsx')
    worksheet = workbook.add_worksheet()

    # Widen the first column to make the text clearer.
    worksheet.set_column('A:A', 30)
    worksheet.set_column('B:B', 30)
    worksheet.set_column('C:C', 30)
    worksheet.set_column('D:D', 30)
    worksheet.set_column('E:E', 30)
    worksheet.set_column('F:F', 30)
    worksheet.set_column('G:G', 30)
    worksheet.set_column('H:H', 30)


    worksheet.write(0, 0, "PUBLISHER")
    worksheet.write(0, 1, "F2F URL")
    worksheet.write(0, 2, "TITLE")
    worksheet.write(0, 3, "SOURCE URL")
    worksheet.write(0, 4, "RECIPE ID")
    worksheet.write(0, 5, "IMAGE URL")
    worksheet.write(0, 6, "SOCIAL RANK")
    worksheet.write(0, 7, "PUBLISHER URL")

    workbook.close()
    """


def insert_ingredient_count():
    file_path = '/home/navendu/Downloads/recipes_351_700.xlsx'
    book = load_workbook(file_path)
    sheet_name = 'recipe_bank'
    sheet = book.get_sheet_by_name(sheet_name)

    row = 1
    for i in sheet.columns[2]:

        if sheet.cell(row=row, column=11).value:
            row += 1
            continue

        print row
        url = i.value
        if 'food2fork' in url:
            resp = urllib2.urlopen(url)
            op = resp.read()

            count = op.count('itemprop="ingredients"')

            items = op.split('<li itemprop="ingredients">')
            items = items[1:]
            ings = []
            for i in items:
                one_ing = i.split('</li>')[0].strip()
                ings.append(one_ing)

            sheet.cell(row=row, column=12).value = count
            sheet.cell(row=row, column=13).value = "\n".join(ings)

            row += 1
            sleep(1)
            if row % 100 == 0:
                book.save(file_path)

    book.save(file_path)




def add_ingredients():
    file_path = '/home/navendu/Downloads/All_Records_v4_9.27_bak.xlsx'
    book = load_workbook(file_path)
    sheet_name = 'All_Records_v4_9.13 (10500 Reco'
    sheet = book.get_sheet_by_name(sheet_name)


    row = 1
    for i in sheet.columns[2]:

        if sheet.cell(row=row, column=13).value:
            row += 1
            continue

        print row
        url = i.value
        if 'food2fork' in url:
            resp = urllib2.urlopen(url)
            op = resp.read()
            items = op.split('<li itemprop="ingredients">')
            items = items[1:]

            ings = []
            for i in items:
                one_ing = i.split('</li>')[0].strip()
                ings.append(one_ing)

            sheet.cell(row=row, column=13).value = "\n".join(ings)
            row += 1
            sleep(1)
            if row % 20 == 0:
                book.save(file_path)

    book.save(file_path)




if __name__ == "__main__":
    #insert_excel()
    #insert_ingredient_count()
    add_ingredients()
